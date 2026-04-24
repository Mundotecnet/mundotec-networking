"""Renderizadores XLSX — MundoTec Network Manager."""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta corporativa ─────────────────────────────────────────────────────────
NAVY    = "0D1B2A"
NARANJA = "E85D04"
LIGHT   = "F3F4F6"
GRAY    = "6B7280"


def _hdr(ws, row, n_cols, bg=NAVY):
    """Aplica estilo de encabezado a una fila."""
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    alin = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = alin


def _title_row(ws, texto, n_cols):
    """Fila de título del reporte."""
    ws.append([texto] + [""] * (n_cols - 1))
    ws.merge_cells(start_row=ws.max_row, start_column=1,
                   end_row=ws.max_row, end_column=n_cols)
    c = ws.cell(row=ws.max_row, column=1)
    c.font = Font(bold=True, size=13, color=NAVY, name="Arial")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[ws.max_row].height = 22


def _meta_row(ws, cliente, fecha, n_cols):
    ws.append([f"Cliente: {cliente}", "", f"Generado: {fecha}"] + [""] * (n_cols - 3))
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9, color=GRAY, name="Arial")
    ws.cell(row=ws.max_row, column=3).font = Font(italic=True, size=9, color=GRAY, name="Arial")
    ws.append([])


def _auto_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, max_w)


def _data_row(ws, values):
    ws.append(values)
    row = ws.max_row
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(vertical="center", wrap_text=False)
    if row % 2 == 0:
        for col in range(1, len(values) + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=LIGHT)


# ══════════════════════════════════════════════════════════════════════════════
# XLSX Infraestructura — multi-sitio
# ══════════════════════════════════════════════════════════════════════════════

def render_xlsx_infraestructura(ctx: dict, out: Path):
    wb = openpyxl.Workbook()
    cliente  = ctx.get("cliente_nombre", "—")
    fecha    = ctx.get("generado_en", "—")
    sitios   = ctx.get("sitios", [])

    # ── Hoja Resumen ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    _title_row(ws, "Mundotec, S.A. — Informe de Infraestructura Actual", 2)
    _meta_row(ws, cliente, fecha, 2)

    total_eq  = sum(len(s.get("equipos", [])) for s in sitios)
    total_vl  = sum(len(s.get("vlans",   [])) for s in sitios)
    total_pp  = sum(s.get("pp_count", 0)      for s in sitios)
    total_lic = len(ctx.get("licencias", []))

    ws.append(["Métrica", "Valor"])
    _hdr(ws, ws.max_row, 2)
    for lbl, val in [
        ("Sitios documentados",       ctx.get("total_sitios", 0)),
        ("Total equipos",             total_eq),
        ("Total VLANs",               total_vl),
        ("Total patch panels",        total_pp),
        ("Licencias registradas",     total_lic),
    ]:
        _data_row(ws, [lbl, val])

    # ── Hoja Equipos ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Equipos")
    cols2 = ["Sitio", "Cuarto", "Nombre", "Categoría", "Tipo",
             "Marca", "Modelo", "Serial", "IP Gestión", "MAC", "Hostname"]
    _title_row(ws2, f"Equipos — {cliente}", len(cols2))
    _meta_row(ws2, cliente, fecha, len(cols2))
    ws2.append(cols2)
    _hdr(ws2, ws2.max_row, len(cols2))

    for s in sitios:
        for eq in s.get("equipos", []):
            _data_row(ws2, [
                s["nombre"],
                eq.get("cuarto_nombre", "—"),
                eq.get("nombre", "—"),
                eq.get("categoria", "—"),
                eq.get("tipo", "—"),
                eq.get("marca", "—"),
                eq.get("modelo", "—"),
                eq.get("serial", "—"),
                eq.get("ip_gestion", "—"),
                eq.get("mac", "—"),
                eq.get("hostname", "—"),
            ])

    ws2.auto_filter.ref = f"A3:{get_column_letter(len(cols2))}{ws2.max_row}"
    _auto_width(ws2)

    # ── Hoja VLANs ────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("VLANs")
    cols3 = ["Sitio", "VLAN ID", "Nombre", "Subred", "Gateway", "DHCP"]
    _title_row(ws3, f"VLANs — {cliente}", len(cols3))
    _meta_row(ws3, cliente, fecha, len(cols3))
    ws3.append(cols3)
    _hdr(ws3, ws3.max_row, len(cols3), NARANJA)

    for s in sitios:
        for v in s.get("vlans", []):
            _data_row(ws3, [
                s["nombre"],
                v.get("vlan_id", "—"),
                v.get("nombre", "—"),
                v.get("cidr", "—"),
                v.get("gateway", "—"),
                "Sí" if v.get("dhcp") else "No",
            ])
    ws3.auto_filter.ref = f"A3:{get_column_letter(len(cols3))}{ws3.max_row}"
    _auto_width(ws3)

    # ── Hoja Patch Panels — resumen + detalle de puertos ─────────────────────
    ws4 = wb.create_sheet("Patch Panels")
    N_COLS4 = 13  # columnas del detalle de puertos

    _title_row(ws4, f"Patch Panels — {cliente}", N_COLS4)
    _meta_row(ws4, cliente, fecha, N_COLS4)

    # Paleta de estados
    COLOR_EST = {
        "completo":    "D1FAE5",  # verde claro
        "parcial":     "FEF3C7",  # amarillo claro
        "sin_revisar": "FEE2E2",  # rojo claro
    }
    FONT_EST = {
        "completo":    "15803D",
        "parcial":     "92400E",
        "sin_revisar": "991B1B",
    }

    cols_resumen = ["Sitio", "Cuarto", "Panel", "Piso", "Marca", "Modelo",
                    "Total", "Documentados", "Sin Revisar", "Completos"]
    cols_puertos = ["Sitio", "Cuarto", "Panel", "Puerto #", "Etiqueta",
                    "Estado", "Tipo", "Descripción / Destino", "VLAN", "IP", "MAC", "Notas", "Completitud"]

    for s in sitios:
        for ed in s.get("edificios", []):
            for cuarto in ed.get("cuartos", []):
                for pp in cuarto.get("paneles", []):
                    # ── Bloque encabezado del panel ──────────────────────────
                    ws4.append([])  # separador

                    # Fila título del panel
                    panel_title = (f"Panel: {pp.get('nombre','—')}   "
                                   f"│  Cuarto: {cuarto.get('nombre','—')}   "
                                   f"│  Sitio: {s['nombre']}   "
                                   f"│  Piso: {pp.get('piso','—')}   "
                                   f"│  {pp.get('marca','—')} {pp.get('modelo','—')}   "
                                   f"│  {pp.get('total_puertos',0)} puertos")
                    ws4.append([panel_title] + [""] * (N_COLS4 - 1))
                    ws4.merge_cells(start_row=ws4.max_row, start_column=1,
                                    end_row=ws4.max_row, end_column=N_COLS4)
                    hdr_cell = ws4.cell(row=ws4.max_row, column=1)
                    hdr_cell.fill = PatternFill("solid", fgColor=NAVY)
                    hdr_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                    hdr_cell.alignment = Alignment(horizontal="left", vertical="center")
                    ws4.row_dimensions[ws4.max_row].height = 16

                    # Fila sub-header columnas
                    ws4.append(cols_puertos)
                    _hdr(ws4, ws4.max_row, N_COLS4, NARANJA)

                    # ── Filas de puertos ─────────────────────────────────────
                    for pt in pp.get("puertos", []):
                        estado = pt.get("estado", "sin_revisar")
                        vals = [
                            s["nombre"],
                            cuarto.get("nombre", "—"),
                            pp.get("nombre", "—"),
                            pt.get("numero", "—"),
                            pt.get("etiqueta", "—"),
                            estado,
                            pt.get("tipo_nodo", "—"),
                            pt.get("descripcion", "—"),
                            pt.get("vlan", "—"),
                            pt.get("ip", "—"),
                            pt.get("mac", "—"),
                            pt.get("notas", "—"),
                            pt.get("completitud", "—"),
                        ]
                        ws4.append(vals)
                        row = ws4.max_row
                        for col in range(1, N_COLS4 + 1):
                            c = ws4.cell(row=row, column=col)
                            c.font = Font(name="Arial", size=9)
                            c.alignment = Alignment(vertical="center", wrap_text=False)
                        # Color según estado en columna "Estado" (col 6)
                        bg = COLOR_EST.get(estado, "FFFFFF")
                        fg = FONT_EST.get(estado, "111827")
                        ec = ws4.cell(row=row, column=6)
                        ec.fill = PatternFill("solid", fgColor=bg)
                        ec.font = Font(name="Arial", size=9, bold=True, color=fg)

    # Anchos fijos por columna (puertos tienen texto largo)
    anchos = [12, 14, 8, 9, 12, 12, 12, 40, 18, 16, 20, 20, 12]
    for i, w in enumerate(anchos, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja Licencias ────────────────────────────────────────────────────────
    licencias = ctx.get("licencias", [])
    if licencias:
        ws5 = wb.create_sheet("Licencias")
        cols5 = ["Producto", "Tipo", "Proveedor", "Vencimiento", "Estado",
                 "Activaciones Máx.", "Activaciones Usadas"]
        _title_row(ws5, f"Licencias — {cliente}", len(cols5))
        _meta_row(ws5, cliente, fecha, len(cols5))
        ws5.append(cols5)
        _hdr(ws5, ws5.max_row, len(cols5))
        for lic in licencias:
            _data_row(ws5, [
                lic.get("producto", "—"), lic.get("tipo", "—"),
                lic.get("proveedor", "—"), lic.get("fecha_vencimiento", "—"),
                lic.get("estado", "—"), lic.get("activaciones_max", "—"),
                lic.get("activaciones_usadas", "—"),
            ])
        _auto_width(ws5)

    wb.save(str(out))


# ══════════════════════════════════════════════════════════════════════════════
# XLSX Inventario (uso general)
# ══════════════════════════════════════════════════════════════════════════════

def render_xlsx(ctx: dict, out: Path):
    wb = openpyxl.Workbook()
    cliente = ctx.get("cliente_nombre", "—")
    fecha   = ctx.get("generado_en",   "—")

    # Hoja Resumen
    ws = wb.active
    ws.title = "Resumen"
    _title_row(ws, "Mundotec, S.A. — Inventario Activo", 2)
    _meta_row(ws, cliente, fecha, 2)
    ws.append(["Métrica", "Valor"])
    _hdr(ws, ws.max_row, 2)
    for k, v in ctx.get("resumen", {}).items():
        _data_row(ws, [k, v])
    _auto_width(ws)

    # Hoja Equipos
    ws2 = wb.create_sheet("Equipos")
    cols2 = ["Nombre", "Tipo", "Marca", "Modelo", "IP Gestión", "Sitio", "Cuarto", "Estado"]
    _title_row(ws2, f"Equipos — {cliente}", len(cols2))
    _meta_row(ws2, cliente, fecha, len(cols2))
    ws2.append(cols2)
    _hdr(ws2, ws2.max_row, len(cols2))
    for eq in ctx.get("equipos", []):
        _data_row(ws2, [eq.get(k, "—") for k in
                        ["nombre","tipo","marca","modelo","ip_gestion","sitio_nombre","cuarto_nombre","estado"]])
    ws2.auto_filter.ref = f"A3:{get_column_letter(len(cols2))}{ws2.max_row}"
    _auto_width(ws2)

    # Hoja Licencias
    ws3 = wb.create_sheet("Licencias")
    cols3 = ["Producto", "Tipo", "Proveedor", "Vencimiento", "Estado", "Act. Máx", "Act. Usadas"]
    _title_row(ws3, f"Licencias — {cliente}", len(cols3))
    _meta_row(ws3, cliente, fecha, len(cols3))
    ws3.append(cols3)
    _hdr(ws3, ws3.max_row, len(cols3), NARANJA)
    for lic in ctx.get("licencias", []):
        _data_row(ws3, [lic.get(k, "—") for k in
                        ["producto","tipo","proveedor","fecha_vencimiento","estado","activaciones_max","activaciones_usadas"]])
    _auto_width(ws3)

    wb.save(str(out))
