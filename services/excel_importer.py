"""
Importador Excel para MundoTec Networking.
Cada hoja = un cuarto. Columnas: PUERTO | DETALLE | MAC | PUERTO SWITCH

v2: además de crear rooms/patch_panels legacy, crea cuarto/gabinete/patch_panel
    y endpoint en el schema nuevo para habilitar trazabilidad.
"""
import re
import uuid
from io import BytesIO
from typing import Any
import openpyxl
from sqlalchemy.orm import Session
from sqlalchemy import text
import models

# ── Type detection ─────────────────────────────────────────────────────────────

_PATTERNS = {
    "camara":    re.compile(r"(cam|camara|camera)", re.I),
    "ap":        re.compile(r"(^ap[- /]|access.pt)", re.I),
    "estacion":  re.compile(r"(estacion|pc\b|maquina|maq|workstation)", re.I),
    "prevista":  re.compile(r"prevista", re.I),
    "libre":     re.compile(r"^libre$", re.I),
    "pbx":       re.compile(r"(pbx|central)", re.I),
    "impresora": re.compile(r"(impresora|printer)", re.I),
    "nvr":       re.compile(r"\bnvr\b", re.I),
    "dvr":       re.compile(r"\bdvr\b", re.I),
}


def _detect_type(detail: str) -> str:
    detail = (detail or "").strip()
    for key, pat in _PATTERNS.items():
        if pat.search(detail):
            return key
    return "otro"


def _clean_mac(raw: str) -> tuple[str, str | None]:
    """Returns (mac, ip_if_found)."""
    raw = (raw or "").strip()
    ip = None
    if "IP:" in raw.upper():
        match = re.search(r"IP:\s*([\d.]+)", raw, re.I)
        if match:
            ip = match.group(1)
        raw = re.sub(r"IP:.*", "", raw, flags=re.I).strip(" ;,")
    return raw, ip


def _clean_sw_port(raw: str) -> str:
    raw = (raw or "").strip()
    if re.match(r"^\d+\.0$", raw):
        return raw.split(".")[0]
    return raw


def _detected_to_endpoint_tipo(detected: str) -> str:
    MAP = {
        "camara": "camara", "ap": "ap", "estacion": "pc",
        "pbx": "otro", "impresora": "impresora", "nvr": "nvr", "dvr": "dvr",
    }
    return MAP.get(detected, "otro")


def _detect_format(label: str) -> str:
    if re.match(r"^[0-9]-[A-Z0-9]+-[A-Z0-9]+-[A-Z0-9]+-[A-Z0-9]+-[0-9]{2}$", label):
        return "extended"
    if re.match(r"^[0-9][A-Z]-[A-Z]-[A-Z][0-9]{2}$", label):
        return "full"
    return "simple"


def _parse_label_meta(label: str, fmt: str) -> dict:
    """Extract floor, building, room_letter, rack_id, panel_letter from label."""
    if fmt == "extended":
        m = re.match(r"^([0-9])-([A-Z0-9]+)-([A-Z0-9]+)-([A-Z0-9]+)-([A-Z0-9]+)-[0-9]{2}$", label)
        if m:
            return {"floor": int(m[1]), "building": m[2], "room_letter": m[3],
                    "rack_id": m[4], "panel_letter": m[5]}
    if fmt == "full":
        m = re.match(r"^([0-9])([A-Z])-([A-Z])-([A-Z])[0-9]{2}$", label)
        if m:
            return {"floor": int(m[1]), "room_letter": m[2], "building": m[3],
                    "panel_letter": m[4], "rack_id": None}
    m = re.match(r"^([0-9])([A-Z])-([A-Z])[0-9]{2}$", label)
    if m:
        return {"floor": int(m[1]), "room_letter": m[2], "panel_letter": m[3],
                "building": None, "rack_id": None}
    return {"floor": 1, "room_letter": "A", "building": None, "rack_id": None, "panel_letter": "A"}


def _generate_label(pp: models.PatchPanel, port_number: int) -> str:
    num_str = f"{port_number:02d}"
    if pp.format == "extended":
        return (f"{pp.floor}-{pp.building or 'A'}-{pp.room_letter}-"
                f"{pp.rack_id or 'A'}-{pp.panel_letter}-{num_str}")
    if pp.format == "full":
        return f"{pp.floor}{pp.room_letter}-{pp.building or 'A'}-{pp.panel_letter}{num_str}"
    return f"{pp.floor}{pp.room_letter}-{pp.panel_letter}{num_str}"


# ── Meta info detection in first 8 rows ───────────────────────────────────────

def _extract_room_meta(ws) -> dict:
    meta: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        row_text = " ".join(str(c) for c in row if c)
        sw_m = re.search(r"switch[:\s]+([\w\s-]+)", row_text, re.I)
        if sw_m:
            meta["switch_model"] = sw_m.group(1).strip()[:100]
        ip_m = re.search(r"\b(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\b", row_text)
        if ip_m and "switch_ip" not in meta:
            meta["switch_ip"] = ip_m.group(1)
        mac_m = re.search(r"([0-9A-Fa-f]{2}[:\-][0-9A-Fa-f]{2}[:\-][0-9A-Fa-f]{2}[:\-][0-9A-Fa-f]{2}[:\-][0-9A-Fa-f]{2}[:\-][0-9A-Fa-f]{2})", row_text)
        if mac_m and "switch_mac" not in meta:
            meta["switch_mac"] = mac_m.group(1)
        ap_m = re.search(r"ap[:\s]+([\w\s-]+)", row_text, re.I)
        if ap_m:
            meta["ap_model"] = ap_m.group(1).strip()[:100]
        loc_m = re.search(r"ubicaci[oó]n[:\s]+(.+)", row_text, re.I)
        if loc_m:
            meta["location"] = loc_m.group(1).strip()[:255]
    return meta


# ── Helpers schema v2 ─────────────────────────────────────────────────────────

def _get_or_create_sitio(db: Session, client_id: int, nombre: str) -> str:
    row = db.execute(text(
        "SELECT id::text FROM sitio WHERE cliente_id=:cid AND nombre=:nom"
    ), {"cid": client_id, "nom": nombre}).scalar()
    if row:
        return row
    sid = str(uuid.uuid4())
    db.execute(text(
        "INSERT INTO sitio (id, cliente_id, nombre) VALUES (:id,:cid,:nom)"
    ), {"id": sid, "cid": client_id, "nom": nombre})
    return sid


def _get_or_create_edificio(db: Session, sitio_id: str) -> str:
    row = db.execute(text(
        "SELECT id::text FROM edificio WHERE sitio_id=:sid AND codigo='A'"
    ), {"sid": sitio_id}).scalar()
    if row:
        return row
    eid = str(uuid.uuid4())
    db.execute(text(
        "INSERT INTO edificio (id, sitio_id, codigo, nombre, piso_default) VALUES (:id,:sid,'A','Principal',1)"
    ), {"id": eid, "sid": sitio_id})
    return eid


def _get_or_create_cuarto(db: Session, edificio_id: str, room_legacy_id: int,
                           codigo: str, nombre: str) -> str:
    row = db.execute(text(
        "SELECT id::text FROM cuarto WHERE room_legacy_id=:rid"
    ), {"rid": room_legacy_id}).scalar()
    if row:
        return row
    cid = str(uuid.uuid4())
    db.execute(text("""
        INSERT INTO cuarto (id, edificio_id, room_legacy_id, piso, codigo, nombre)
        VALUES (:id, :eid, :rid, 1, :cod, :nom)
        ON CONFLICT DO NOTHING
    """), {"id": cid, "eid": edificio_id, "rid": room_legacy_id, "cod": codigo, "nom": nombre})
    # re-fetch in case of conflict
    row = db.execute(text("SELECT id::text FROM cuarto WHERE room_legacy_id=:rid"),
                     {"rid": room_legacy_id}).scalar()
    return row or cid


def _create_patch_panel_v2(db: Session, gabinete_id: str, codigo: str,
                            puertos_total: int) -> str:
    ppid = str(uuid.uuid4())
    db.execute(text("""
        INSERT INTO patch_panel (id, gabinete_id, codigo, tipo, categoria, puertos_total)
        VALUES (:id, :gid, :cod, 'cobre', 'Cat6', :total)
        ON CONFLICT DO NOTHING
    """), {"id": ppid, "gid": gabinete_id, "cod": codigo, "total": puertos_total})
    row = db.execute(text(
        "SELECT id::text FROM patch_panel WHERE gabinete_id=:gid AND codigo=:cod"
    ), {"gid": gabinete_id, "cod": codigo}).scalar()
    return row or ppid


def _create_gabinete(db: Session, cuarto_id: str, codigo: str = "A") -> str:
    row = db.execute(text(
        "SELECT id::text FROM gabinete WHERE cuarto_id=:cid AND codigo=:cod"
    ), {"cid": cuarto_id, "cod": codigo}).scalar()
    if row:
        return row
    gid = str(uuid.uuid4())
    db.execute(text(
        "INSERT INTO gabinete (id, cuarto_id, codigo, nombre) VALUES (:id,:cid,:cod,'Gabinete Principal')"
    ), {"id": gid, "cid": cuarto_id, "cod": codigo})
    return gid


def _create_endpoint_v2(db: Session, client_id: int, sitio_id: str,
                         nombre: str, tipo: str, ip: str | None, mac: str | None,
                         faceplate_puerto_id: str | None) -> str:
    eid = str(uuid.uuid4())
    db.execute(text("""
        INSERT INTO endpoint (id, cliente_id, sitio_id, tipo, nombre, ip, mac, faceplate_puerto_id)
        VALUES (:id, :cid, :sid, :tipo, :nom, :ip, :mac, :face)
        ON CONFLICT DO NOTHING
    """), {
        "id": eid, "cid": client_id, "sid": sitio_id,
        "tipo": tipo, "nom": nombre[:200],
        "ip": ip, "mac": mac or None,
        "face": faceplate_puerto_id,
    })
    return eid


# ── Port row parser (shared) ──────────────────────────────────────────────────

def _parse_sheet_ports(ws) -> tuple[list[dict], str, str]:
    """Return (port_data, label_format, first_label). port_data is list of dicts."""
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=15):
        for cell in row:
            v = str(cell.value).strip().upper() if cell.value else ""
            if v in ("PUERTO", "PP") or "PUERTO" in v:
                header_row_idx = cell.row
                break
        if header_row_idx:
            break
    if not header_row_idx:
        return [], "simple", ""

    headers = [str(c.value).strip().upper() if c.value else "" for c in ws[header_row_idx]]
    col_puerto = next((i for i, h in enumerate(headers) if h in ("PUERTO", "PP") or "PUERTO" in h), None)
    col_detalle = next((i for i, h in enumerate(headers) if "DETALLE" in h), None)
    col_mac    = next((i for i, h in enumerate(headers) if "MAC" in h), None)
    col_sw     = next((i for i, h in enumerate(headers) if "SWITCH" in h or "SW" in h), None)

    port_data: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(c is None for c in row):
            continue
        entry: dict = {}
        if col_puerto is not None and col_puerto < len(row):
            entry["port_label"] = str(row[col_puerto] or "").strip()
        if col_detalle is not None and col_detalle < len(row):
            entry["detail"] = str(row[col_detalle] or "").strip()
        if col_mac is not None and col_mac < len(row):
            raw_mac = str(row[col_mac] or "").strip()
            mac, ip = _clean_mac(raw_mac)
            entry["mac"] = mac
            entry["ip_from_mac"] = ip
        else:
            entry["mac"] = ""
            entry["ip_from_mac"] = None
        if col_sw is not None and col_sw < len(row):
            entry["sw_port"] = _clean_sw_port(str(row[col_sw] or ""))
        if entry.get("port_label") or entry.get("detail"):
            port_data.append(entry)

    first_label = next((e["port_label"] for e in port_data if e.get("port_label")), "")
    label_format = _detect_format(first_label) if first_label else "simple"
    return port_data, label_format, first_label


# ── Preview (no DB) ────────────────────────────────────────────────────────────

def preview_excel(file_content: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    sheets = []
    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        port_data, label_format, first_label = _parse_sheet_ports(ws)
        if not port_data:
            sheets.append({"index": idx, "name": sheet_name,
                            "label_format": None, "port_count": 0,
                            "ports": [], "error": "No se encontró columna PUERTO/PP o no hay datos"})
            continue
        preview_ports = []
        for i, e in enumerate(port_data[:24]):
            preview_ports.append({
                "number": i + 1,
                "label": e.get("port_label", ""),
                "detail": e.get("detail", ""),
                "mac": e.get("mac", "") or "",
                "ip": e.get("ip_from_mac", "") or "",
            })
        sheets.append({
            "index": idx,
            "name": sheet_name,
            "label_format": label_format,
            "first_label": first_label,
            "port_count": len(port_data),
            "ports": preview_ports,
        })
    return {"sheets": sheets}


# ── Verify (dry-run) ──────────────────────────────────────────────────────────

def verify_import(
    db: Session,
    file_content: bytes,
    patch_panel_id: int,
    sheet_index: int = 0,
) -> dict:
    pp = db.query(models.PatchPanel).filter(models.PatchPanel.id == patch_panel_id).first()
    if not pp:
        raise ValueError("Patch panel no encontrado")

    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    if sheet_index >= len(wb.sheetnames):
        raise ValueError("Índice de hoja fuera de rango")

    sheet_name = wb.sheetnames[sheet_index]
    ws = wb[sheet_name]
    port_data, _, _ = _parse_sheet_ports(ws)
    if not port_data:
        raise ValueError(f"Hoja '{sheet_name}': no se encontraron datos de puertos")

    ports_by_number = {p.number: p for p in pp.ports}
    panel_port_count = len(pp.ports)
    excel_row_count = len(port_data)

    if excel_row_count < panel_port_count:
        # warn but continue — missing rows leave ports unchanged
        pass

    changes = []
    for number in sorted(ports_by_number.keys()):
        port = ports_by_number[number]
        idx = number - 1
        if idx >= excel_row_count:
            changes.append({
                "port_number": number,
                "port_label": port.label or "",
                "current": {
                    "description": port.node_description or "",
                    "mac": port.node_mac or "",
                    "ip": port.node_ip or "",
                    "status": port.completeness_status or "",
                },
                "new": None,
                "will_change": False,
                "skipped": True,
            })
            continue

        e = port_data[idx]
        detail = e.get("detail", "")
        detected = _detect_type(detail)

        if detected in ("libre", "prevista"):
            new_state = {"description": "", "mac": "", "ip": "", "status": detected}
        else:
            new_state = {
                "description": detail,
                "mac": e.get("mac") or "",
                "ip": e.get("ip_from_mac") or "",
                "status": "sin_revisar",
            }

        current_state = {
            "description": port.node_description or "",
            "mac": port.node_mac or "",
            "ip": port.node_ip or "",
            "status": port.completeness_status or "",
        }

        will_change = current_state != new_state
        changes.append({
            "port_number": number,
            "port_label": e.get("port_label") or port.label or "",
            "current": current_state,
            "new": new_state,
            "will_change": will_change,
            "skipped": False,
        })

    changed_count = sum(1 for c in changes if c["will_change"])
    return {
        "panel_id": pp.id,
        "panel_name": pp.name,
        "sheet_name": sheet_name,
        "panel_port_count": panel_port_count,
        "excel_row_count": excel_row_count,
        "will_change": changed_count,
        "unchanged": len([c for c in changes if not c["will_change"] and not c.get("skipped")]),
        "skipped": len([c for c in changes if c.get("skipped")]),
        "ports": changes,
    }


# ── Import into existing panel ─────────────────────────────────────────────────

def import_into_panel(
    db: Session,
    file_content: bytes,
    patch_panel_id: int,
    sheet_index: int = 0,
    current_user=None,
) -> dict:
    pp = db.query(models.PatchPanel).filter(models.PatchPanel.id == patch_panel_id).first()
    if not pp:
        raise ValueError("Patch panel no encontrado")

    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    if sheet_index >= len(wb.sheetnames):
        raise ValueError("Índice de hoja fuera de rango")

    sheet_name = wb.sheetnames[sheet_index]
    ws = wb[sheet_name]
    port_data, _, _ = _parse_sheet_ports(ws)
    if not port_data:
        raise ValueError(f"Hoja '{sheet_name}': no se encontraron datos de puertos")

    ports_updated = 0
    for port in pp.ports:
        idx = port.number - 1
        if idx >= len(port_data):
            continue
        e = port_data[idx]
        detail = e.get("detail", "")
        detected = _detect_type(detail)

        if e.get("port_label"):
            port.label = e["port_label"]

        if detected == "libre":
            port.node_type = "libre"
            port.completeness_status = "libre"
            port.node_description = None
            port.node_mac = None
            port.node_ip = None
        elif detected == "prevista":
            port.node_type = "prevista"
            port.completeness_status = "prevista"
            port.node_description = None
        else:
            if detail:
                port.node_type = "descripcion"
                port.node_description = detail
            port.node_mac = e.get("mac") or None
            port.node_ip = e.get("ip_from_mac") or None
            port.completeness_status = "sin_revisar"

        ports_updated += 1

    db.commit()
    return {
        "panel_id": pp.id,
        "panel_name": pp.name,
        "room_id": pp.room_id,
        "client_id": pp.room.client_id,
        "cabinet_id": pp.cabinet_id,
        "sheet_name": sheet_name,
        "ports_updated": ports_updated,
    }


# ── Main importer ──────────────────────────────────────────────────────────────

def import_excel(
    db: Session,
    file_content: bytes,
    client_name: str,
    current_user: models.User | None = None,
) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    rooms_created = 0
    ports_imported = 0
    warnings: list[str] = []

    # Find or create client
    client = db.query(models.Client).filter(
        models.Client.name.ilike(client_name)
    ).first()
    if not client:
        client = models.Client(name=client_name)
        db.add(client)
        db.flush()

    # Sitio v2 — uno por cliente/importación
    sitio_id = _get_or_create_sitio(db, client.id, client.name)
    edificio_id = _get_or_create_edificio(db, sitio_id)
    cuarto_letra = 65  # 'A'

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        meta = _extract_room_meta(ws)

        room = db.query(models.Room).filter(
            models.Room.client_id == client.id,
            models.Room.name == sheet_name,
        ).first()
        if not room:
            room = models.Room(
                client_id=client.id,
                name=sheet_name,
                location=meta.get("location", ""),
                switch_model=meta.get("switch_model"),
                switch_mac=meta.get("switch_mac"),
                switch_ip=meta.get("switch_ip"),
                ap_model=meta.get("ap_model"),
            )
            db.add(room)
            db.flush()
            rooms_created += 1

        # Cuarto v2
        codigo_cuarto = chr(cuarto_letra) if cuarto_letra <= 90 else str(cuarto_letra - 64)
        cuarto_id = _get_or_create_cuarto(db, edificio_id, room.id, codigo_cuarto, sheet_name)
        gabinete_id = _create_gabinete(db, cuarto_id)
        cuarto_letra += 1

        # Detect headers row (find row containing "PUERTO" or "PP")
        header_row_idx = None
        for row in ws.iter_rows(min_row=1, max_row=15):
            for cell in row:
                v = str(cell.value).strip().upper() if cell.value else ""
                if v in ("PUERTO", "PP") or "PUERTO" in v:
                    header_row_idx = cell.row
                    break
            if header_row_idx:
                break

        if not header_row_idx:
            warnings.append(f"Hoja '{sheet_name}': no se encontró fila de encabezados")
            continue

        headers = [str(c.value).strip().upper() if c.value else "" for c in ws[header_row_idx]]
        col_puerto = next((i for i, h in enumerate(headers) if h in ("PUERTO", "PP") or "PUERTO" in h), None)
        col_detalle = next((i for i, h in enumerate(headers) if "DETALLE" in h), None)
        col_mac = next((i for i, h in enumerate(headers) if "MAC" in h), None)
        col_sw = next((i for i, h in enumerate(headers) if "SWITCH" in h or "SW" in h), None)

        if col_puerto is None:
            warnings.append(f"Hoja '{sheet_name}': columna PUERTO/PP no encontrada")
            continue

        # Group ports into patches of 24
        port_data: list[dict] = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(c is None for c in row):
                continue
            entry: dict = {}
            if col_puerto is not None and col_puerto < len(row):
                entry["port_label"] = str(row[col_puerto] or "").strip()
            if col_detalle is not None and col_detalle < len(row):
                entry["detail"] = str(row[col_detalle] or "").strip()
            if col_mac is not None and col_mac < len(row):
                raw_mac = str(row[col_mac] or "").strip()
                mac, ip = _clean_mac(raw_mac)
                entry["mac"] = mac
                entry["ip_from_mac"] = ip
            else:
                entry["mac"] = ""
                entry["ip_from_mac"] = None
            if col_sw is not None and col_sw < len(row):
                entry["sw_port"] = _clean_sw_port(str(row[col_sw] or ""))
            if entry.get("port_label") or entry.get("detail"):
                port_data.append(entry)

        if not port_data:
            warnings.append(f"Hoja '{sheet_name}': sin datos de puertos")
            continue

        # Determine label format from first labeled port
        label_format = "simple"
        first_label = next((e["port_label"] for e in port_data if e.get("port_label")), "")
        if first_label:
            label_format = _detect_format(first_label)

        # Parse PP info from first label
        meta_pp = _parse_label_meta(first_label, label_format)
        floor = meta_pp["floor"]
        room_letter = meta_pp["room_letter"]
        building = meta_pp["building"]
        panel_letter = meta_pp["panel_letter"]
        rack_id = meta_pp.get("rack_id")

        # Create patch panel (legacy)
        pp = models.PatchPanel(
            room_id=room.id,
            name=f"PP-{panel_letter}",
            floor=floor,
            building=building,
            room_letter=room_letter,
            panel_letter=panel_letter,
            rack_id=rack_id,
            format=label_format,
        )
        db.add(pp)
        db.flush()

        # Patch panel v2
        pp_v2_id = _create_patch_panel_v2(db, gabinete_id, panel_letter, 24)

        # Create 24 ports
        for i in range(1, 25):
            label = _generate_label(pp, i)
            idx = i - 1
            if idx < len(port_data):
                e = port_data[idx]
                detail = e.get("detail", "")
                detected = _detect_type(detail)
                if detected == "libre":
                    node_type, status = "libre", "libre"
                elif detected == "prevista":
                    node_type, status = "prevista", "prevista"
                else:
                    node_type, status = "descripcion", "sin_revisar"

                port = models.PatchPort(
                    patch_panel_id=pp.id,
                    number=i,
                    label=e.get("port_label") or label,
                    node_type=node_type if detected not in ("libre", "prevista") else detected,
                    node_description=detail if detected not in ("libre", "prevista") else None,
                    node_mac=e.get("mac") or None,
                    node_ip=e.get("ip_from_mac") or None,
                    status=status,
                    completeness_status=status,
                )

                # Puerto terminal v2
                etq_norm = f"PP-1-A-{codigo_cuarto}-A-{panel_letter}-{i:02d}"
                port_v2_id = str(uuid.uuid4())
                db.execute(text("""
                    INSERT INTO puerto_terminal
                      (id, tipo, patch_panel_id, numero, etiqueta_norm, etiqueta_display, notas)
                    VALUES (:id,'patch_panel_port',:pp,:num,:etq,:disp,:notes)
                    ON CONFLICT (etiqueta_norm) DO NOTHING
                """), {
                    "id": port_v2_id, "pp": pp_v2_id, "num": i,
                    "etq": etq_norm,
                    "disp": e.get("port_label") or label,
                    "notes": detail if detail else None,
                })

                # Endpoint v2 si tiene datos útiles
                if detected not in ("libre", "prevista") and detail:
                    tipo_ep = _detected_to_endpoint_tipo(detected)
                    nombre_ep = f"{detail[:80]} [{label}]"
                    _create_endpoint_v2(
                        db, client.id, sitio_id,
                        nombre_ep, tipo_ep,
                        e.get("ip_from_mac"), e.get("mac"),
                        port_v2_id,
                    )
            else:
                port = models.PatchPort(
                    patch_panel_id=pp.id,
                    number=i,
                    label=label,
                    status="sin_revisar",
                    completeness_status="sin_revisar",
                )
            db.add(port)
            ports_imported += 1

    db.commit()
    return {
        "client_id": client.id,
        "rooms_created": rooms_created,
        "ports_imported": ports_imported,
        "warnings": warnings,
    }
